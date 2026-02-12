#!/usr/bin/env python3
"""
Aggregate MF data by Theme + FundFamily
Sums tv_ columns for all stocks within each theme
Calculates bb_ (bucket bands) from tv_ values
"""

from pathlib import Path
import pandas as pd
import numpy as np
from app import build_theme_map, DATA_PATH_DEFAULT
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from mf_processor import find_latest_pivot_file


def tv_to_bb(tv):
    """
    Convert total value (tv) to bucket band (bb)

    Extended bucketing logic with -5 to 5 range:
    - tv >= 1000:   bb = 5   (very strong buying)
    - tv >= 100:    bb = 4   (strong buying)
    - tv >= 50:     bb = 3   (moderate-strong buying)
    - tv >= 10:     bb = 2   (moderate buying)
    - tv > 0:       bb = 1   (light buying)
    - tv == 0:      bb = 0   (neutral)
    - tv >= -10:    bb = -1  (light selling)
    - tv >= -50:    bb = -2  (moderate selling)
    - tv >= -100:   bb = -3  (moderate-strong selling)
    - tv >= -1000:  bb = -4  (strong selling)
    - tv < -1000:   bb = -5  (very strong selling)
    """
    if pd.isna(tv):
        return np.nan

    if tv >= 1000:
        return 5
    elif tv >= 100:
        return 4
    elif tv >= 50:
        return 3
    elif tv >= 10:
        return 2
    elif tv > 0:
        return 1
    elif tv == 0:
        return 0
    elif tv >= -10:
        return -1
    elif tv >= -50:
        return -2
    elif tv >= -100:
        return -3
    elif tv >= -1000:
        return -4
    else:
        return -5


def format_aggregated_sheet(ws, df, portfolio_themes):
    """Apply formatting to Aggregated sheet with alternating colors per theme"""

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Alternating theme colors
    theme_fill_1 = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")  # Light blue
    theme_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

    # Portfolio theme highlight (light yellow)
    portfolio_fill_1 = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    portfolio_fill_2 = PatternFill(start_color="FFFAEB", end_color="FFFAEB", fill_type="solid")

    # TOTAL row highlight
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True, size=11)

    # Borders
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Format header row
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Track current theme for alternating colors
    current_theme = None
    theme_color_index = 0

    # Format data rows
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        theme = row_data[0]  # Theme column
        fund_family = row_data[1]  # FundFamily column

        # Change color when theme changes
        if theme != current_theme:
            current_theme = theme
            theme_color_index = 1 - theme_color_index  # Toggle between 0 and 1

        # Determine fill color
        is_total_row = fund_family == "TOTAL"
        is_portfolio_theme = theme in portfolio_themes

        if is_total_row:
            fill = total_fill
            font = total_font
        elif is_portfolio_theme:
            fill = portfolio_fill_1 if theme_color_index == 0 else portfolio_fill_2
            font = Font(size=10)
        else:
            fill = theme_fill_1 if theme_color_index == 0 else theme_fill_2
            font = Font(size=10)

        # Apply formatting to all cells in the row
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.font = font
            cell.border = thin_border

            # Align numbers to the right
            if col_num > 2:  # Skip Theme and FundFamily columns
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Set column widths
    ws.column_dimensions['A'].width = 18  # Theme
    ws.column_dimensions['B'].width = 15  # FundFamily

    # Set widths for tv_ and bb_ columns differently
    # Columns 3 onwards: tv_MonYY, tv_MonYY, ..., bb_MonYY, bb_MonYY, ...
    num_tv_cols = (len(df.columns) - 2) // 2  # Number of tv_ columns

    for col_num in range(3, len(df.columns) + 1):
        col_letter = get_column_letter(col_num)
        if col_num <= 2 + num_tv_cols:
            # tv_ columns - standard width
            ws.column_dimensions[col_letter].width = 12
        else:
            # bb_ columns - narrower width for visual separation
            ws.column_dimensions[col_letter].width = 6

    # Freeze header row and first two columns
    ws.freeze_panes = 'C2'


def format_debug_sheet(ws, df):
    """Apply formatting to Debug sheet with alternating colors per symbol"""

    # Define styles
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Alternating symbol colors
    symbol_fill_1 = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green
    symbol_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

    # Portfolio symbol highlight
    portfolio_fill_1 = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    portfolio_fill_2 = PatternFill(start_color="FFFAEB", end_color="FFFAEB", fill_type="solid")

    # Borders
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Format header row
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Track current symbol for alternating colors
    current_symbol = None
    symbol_color_index = 0

    # Format data rows
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        symbol = row_data[1]  # Symbol column
        is_portfolio = row_data[3]  # IsPortfolio column

        # Change color when symbol changes
        if symbol != current_symbol:
            current_symbol = symbol
            symbol_color_index = 1 - symbol_color_index  # Toggle between 0 and 1

        # Determine fill color
        if is_portfolio:
            fill = portfolio_fill_1 if symbol_color_index == 0 else portfolio_fill_2
        else:
            fill = symbol_fill_1 if symbol_color_index == 0 else symbol_fill_2

        # Apply formatting to all cells in the row
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.font = Font(size=10)
            cell.border = thin_border

            # Align numbers to the right, text to the left
            if col_num > 4:  # Numeric columns
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Set column widths
    ws.column_dimensions['A'].width = 18  # Theme
    ws.column_dimensions['B'].width = 15  # Symbol
    ws.column_dimensions['C'].width = 15  # FundFamily
    ws.column_dimensions['D'].width = 12  # IsPortfolio

    # Set widths for tv_ and bb_ columns differently
    # Columns 5 onwards: tv_MonYY, ..., bb_MonYY, ...
    num_tv_cols = (len(df.columns) - 4) // 2  # Number of tv_ columns

    for col_num in range(5, len(df.columns) + 1):
        col_letter = get_column_letter(col_num)
        if col_num <= 4 + num_tv_cols:
            # tv_ columns - standard width
            ws.column_dimensions[col_letter].width = 12
        else:
            # bb_ columns - narrower width for visual separation
            ws.column_dimensions[col_letter].width = 6

    # Freeze header row and first four columns
    ws.freeze_panes = 'E2'


def aggregate_by_theme_and_fund():
    """
    Aggregate MF data by Theme + FundFamily
    Returns: (aggregated_df, debug_df)
    """

    # Load theme mapping
    print("Loading theme map...")
    th = pd.read_excel(DATA_PATH_DEFAULT, sheet_name="theme_park")
    theme_map = build_theme_map(th)

    # Load portfolio symbols to identify portfolio themes
    pf = pd.read_excel(DATA_PATH_DEFAULT, sheet_name="PF_Ranks")
    pf = pf.rename(columns={"Symbol / Rank": "Symbol"})
    pf["Symbol"] = pf["Symbol"].astype(str).str.strip()
    portfolio_symbols = set(pf["Symbol"].dropna())

    # Identify portfolio themes
    portfolio_themes = sorted(theme_map[theme_map['Symbol'].isin(portfolio_symbols)]['Theme'].unique())

    # Load MF data
    print("Loading MF data...")
    mf_file, mf_date_label = find_latest_pivot_file()
    print(f"Using pivot file: {mf_file.name} ({mf_date_label})")
    mf_df = pd.read_excel(mf_file, sheet_name="Summary Data")

    # Merge with theme map
    print("Merging with theme map...")
    mf_with_theme = mf_df.merge(theme_map, on='Symbol', how='left')

    # Filter only rows with theme mapping
    mf_with_theme = mf_with_theme[mf_with_theme['Theme'].notna()]

    # Add IsPortfolio flag
    mf_with_theme['IsPortfolio'] = mf_with_theme['Symbol'].isin(portfolio_symbols)

    # Get all tv_ columns
    tv_cols = [c for c in mf_with_theme.columns if c.startswith('tv_')]

    print(f"Aggregating {len(tv_cols)} tv_ columns: {tv_cols}")

    # Create debug dataframe BEFORE aggregation
    print("Creating debug tab with stock-level details...")
    debug_cols = ['Theme', 'Symbol', 'FundFamily', 'IsPortfolio'] + tv_cols
    debug_df = mf_with_theme[debug_cols].copy()

    # Round tv_ columns to integers in debug tab
    for tv_col in tv_cols:
        debug_df[tv_col] = debug_df[tv_col].round(0).astype('Int64')

    # Calculate bb_ columns for debug tab (per-symbol, from original MF data)
    # BB values are per-symbol (same across all fund families), so we get them from the input
    print("Adding bb_ columns from source data (per-symbol)...")
    for tv_col in tv_cols:
        bb_col = tv_col.replace('tv_', 'bb_')
        # Get bb_ directly from the original MF data (already calculated per-symbol)
        if bb_col in mf_with_theme.columns:
            debug_df[bb_col] = mf_with_theme[bb_col].astype('Int64')
        else:
            # Fallback if bb_ column doesn't exist
            debug_df[bb_col] = debug_df[tv_col].apply(tv_to_bb).astype('Int64')

    # Sort debug df: Portfolio themes first, then by Theme, Symbol, FundFamily
    debug_df = debug_df.sort_values(
        by=['IsPortfolio', 'Theme', 'Symbol', 'FundFamily'],
        ascending=[False, True, True, True]
    )

    # Reorder debug columns: Theme, Symbol, FundFamily, IsPortfolio, all tv_ columns, then all bb_ columns
    bb_cols_debug = [tv_col.replace('tv_', 'bb_') for tv_col in tv_cols]
    debug_final_cols = ['Theme', 'Symbol', 'FundFamily', 'IsPortfolio'] + tv_cols + bb_cols_debug
    debug_df = debug_df[debug_final_cols]

    # Group by Theme + FundFamily and sum tv_ columns
    print("Aggregating by Theme + FundFamily...")
    agg_dict = {col: 'sum' for col in tv_cols}

    theme_fund_agg = mf_with_theme.groupby(['Theme', 'FundFamily'], as_index=False).agg(agg_dict)

    # Sort: Portfolio themes first, then others (alphabetically within each group)
    print("Sorting themes (portfolio first)...")
    theme_fund_agg['IsPortfolio'] = theme_fund_agg['Theme'].isin(portfolio_themes)
    theme_fund_agg = theme_fund_agg.sort_values(
        by=['IsPortfolio', 'Theme', 'FundFamily'],
        ascending=[False, True, True]
    )

    # Add Total rows for each theme and sort fund families by latest TV
    print("Adding Total rows for each theme...")
    print("Sorting fund families by latest TV (highest first)...")
    result_rows = []

    # Get the latest tv_ column (last one chronologically)
    latest_tv_col = tv_cols[-1]  # e.g., tv_Jan26

    for theme in theme_fund_agg['Theme'].unique():
        theme_data = theme_fund_agg[theme_fund_agg['Theme'] == theme]

        # Sort fund families by latest TV value (descending - highest first)
        theme_data_sorted = theme_data.sort_values(by=latest_tv_col, ascending=False)

        # Add all fund family rows for this theme
        for _, row in theme_data_sorted.iterrows():
            result_rows.append(row.to_dict())

        # Add Total row at the end
        total_row = {'Theme': theme, 'FundFamily': 'TOTAL', 'IsPortfolio': theme in portfolio_themes}
        for col in tv_cols:
            total_row[col] = theme_data[col].sum()
        result_rows.append(total_row)

    # Create final dataframe
    result_df = pd.DataFrame(result_rows)

    # Drop IsPortfolio helper column
    result_df = result_df.drop(columns=['IsPortfolio'])

    # Round tv_ columns to integers
    print("Rounding tv_ values to integers...")
    for tv_col in tv_cols:
        result_df[tv_col] = result_df[tv_col].round(0).astype('Int64')

    # Calculate bb_ columns from TOTAL tv_ values (theme-level)
    print("Calculating bb_ (bucket bands) from TOTAL tv_ values per theme...")

    # First, get TOTAL rows only to calculate bb_ values
    total_rows = result_df[result_df['FundFamily'] == 'TOTAL'].copy()

    # Calculate bb_ for each theme based on TOTAL tv_
    theme_bb_map = {}
    for _, row in total_rows.iterrows():
        theme = row['Theme']
        theme_bb_map[theme] = {}
        for tv_col in tv_cols:
            bb_col = tv_col.replace('tv_', 'bb_')
            theme_bb_map[theme][bb_col] = tv_to_bb(row[tv_col])

    # Apply bb_ values to all rows of each theme
    bb_cols = []
    for tv_col in tv_cols:
        bb_col = tv_col.replace('tv_', 'bb_')
        result_df[bb_col] = result_df.apply(
            lambda row: theme_bb_map[row['Theme']][bb_col], axis=1
        )
        result_df[bb_col] = result_df[bb_col].astype('Int64')
        bb_cols.append(bb_col)

    # Reorder columns: Theme, FundFamily, all tv_ columns, then all bb_ columns
    cols = ['Theme', 'FundFamily'] + tv_cols + bb_cols
    result_df = result_df[cols]

    # Save to Excel with 2 tabs
    output_file = Path(f"/Users/raviaranke/Desktop/themes/{mf_date_label}_theme_aggregated.xlsx")
    print(f"\nSaving to {output_file}...")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Tab 1: Aggregated output
        result_df.to_excel(writer, sheet_name="Aggregated", index=False)

        # Tab 2: Debug/Audit trail
        debug_df.to_excel(writer, sheet_name="Debug", index=False)

        # Apply formatting
        print("Applying Excel formatting...")
        format_aggregated_sheet(writer.sheets["Aggregated"], result_df, portfolio_themes)
        format_debug_sheet(writer.sheets["Debug"], debug_df)

    print(f"✅ Done! Created {output_file}")
    print(f"\n📊 Tab 1 (Aggregated):")
    print(f"   - {len(result_df)} rows")
    print(f"   - {len(result_df['Theme'].unique())} unique themes")
    print(f"   - {len(portfolio_themes)} portfolio themes")
    print(f"   - {len(tv_cols)} months of data (tv_ and bb_ columns)")
    print(f"   - Total columns: {len(cols)}")

    print(f"\n🔍 Tab 2 (Debug):")
    print(f"   - {len(debug_df)} rows (stock-level details)")
    print(f"   - {len(debug_df['Symbol'].unique())} unique symbols")
    print(f"   - Columns: Theme, Symbol, FundFamily, IsPortfolio, all tv_ cols, then all bb_ cols")
    print(f"   - BB values are per-symbol (same across all fund families for each symbol)")
    print(f"   - Use Excel filters to verify: pick a theme, see all stocks & sums")

    return result_df, debug_df


if __name__ == "__main__":
    agg_df, debug_df = aggregate_by_theme_and_fund()

    # Show sample from aggregated tab
    print("\n" + "="*80)
    print("SAMPLE - Tab 1 (Aggregated) - First 10 rows:")
    print("="*80)
    # Get last 2 months of tv and bb columns for display
    tv_cols_display = [c for c in agg_df.columns if c.startswith('tv_')][-2:]
    bb_cols_display = [c for c in agg_df.columns if c.startswith('bb_')][-2:]
    sample_cols = ['Theme', 'FundFamily'] + tv_cols_display + bb_cols_display
    print(agg_df[sample_cols].head(10).to_string(index=False))
    print("\nNote: BB values are per-theme (same for all fund families within each theme)")

    # Show sample from debug tab
    print("\n" + "="*80)
    print("SAMPLE - Tab 2 (Debug) - First 10 rows:")
    print("="*80)
    debug_sample_cols = ['Theme', 'Symbol', 'FundFamily', 'IsPortfolio'] + tv_cols_display + bb_cols_display
    print(debug_df[debug_sample_cols].head(10).to_string(index=False))
    print("\nNote: BB values are per-symbol (same for all fund families of each symbol)")
